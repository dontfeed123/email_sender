import openpyxl
from openpyxl.utils.exceptions import InvalidFileException, SheetTitleException


def load_workbook_and_get_emails(path_to_wb: str, sheet_name: str, column_id: str,column_emails: str) -> list:
    """
        This function open a xlsx file and returns all the emails in a list

    :param path_to_wb: This is the path to the xlsx file. E.g: /home/mendes/workbook.xlsx
    :param sheet_name: This is the workbook sheet name, which contains the emails data
    :param column_emails: This is a letter that represent the letter of the column. E.g: A, B, C, D
    :return: This function returns a list of emails.
    """

    try:
        sheet = openpyxl.load_workbook(path_to_wb)[sheet_name]
    except KeyError:
        raise KeyError("Any sheet found with that title.")
    except InvalidFileException:
        raise InvalidFileException("Couldn't open the path expecified.")
    except SheetTitleException:
        raise SheetTitleException("Sheet title Misspelled.")

    # All rows in the xlsx file
    list_id_with_emails_temp = []
    for row_count in range(2, sheet.max_row + 1):
        try:
            email_number_conversor_from_letter = alphabet_position(column_emails.lower())
            id_number_conversor_from_letter = alphabet_position(column_id.lower())
            _id = sheet.cell(row=row_count, column=int(id_number_conversor_from_letter)).value
            email = sheet.cell(row=row_count, column=int(email_number_conversor_from_letter)).value
            list_id_with_emails_temp.append((_id, email))
        except ValueError:
            print("Value not found in row: {}".format(row_count))
            continue
    list_id_with_emails = []
    for values in list_id_with_emails_temp:
        if values[0] != None and values[1] != None:
            list_id_with_emails.append(values)

    return list_id_with_emails
def alphabet_position(letter: str) -> int:
    """
        Returns the index of a letter passed to the function.
        This used only for openpyxl column values
    :param letter: A letter to convert to int
    :return: Return a int that is the index of the letter in the alphabet
    """

    alphabet = {'a': '1', 'b': '2', 'c': '3', 'd': '4', 'e': '5', 'f': '6', 'g': '7', 'h': '8',
                'i': '9', 'j': '10', 'k': '11', 'l': '12', 'm': '13', 'n': '14', 'o': '15', 'p': '16', 'q': '17',
                'r': '18', 's': '19', 't': '20', 'u': '21', 'v': '22', 'w': '23', 'x': '24', 'y': '25', 'z': '26'
                }
    for key, value in alphabet.items():
        if key == letter:
            return value



print(load_workbook_and_get_emails("./emails.xlsx","Folha1", "B","E"))


