"""
    @version 001:
    Added a conversor to convert column letters in numberrs to easily access the xlsx file in the right column

    Already Have:
        - Send emails to a specification list of emails
    Todo:
        - Errors Handling (important)
        - Multi-Threading
        - Log File
        - Not tested with a big amount of emails

    Functions:
        -alphabet_position() -> Convert a letter in a number
        -email_sender() -> Send emails
        -load_workbook_and_get_emails() -> Open a xlsx and store all the emails.
"""
import smtplib
from email.mime.text import MIMEText
from multiprocessing import Process

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException, SheetTitleException


class GlobalStorage:
    """
        todo The Server Is Set Here
        This class stores all global variables, to all functions access it easily
    """

    def __init__(self, email="testeautomateemails@gmail.com", password="testeTeste21-06-2018"):
        self.my_email = email  # This is to login in smpt server (needs to be gmail)
        self.my_password = password  # This is to login in smpt server

    def connect_to_smtp_server(self):
        # Connection to the server
        self.server = smtplib.SMTP("smtp.gmail.com", 587)
        self.server.starttls()
        self.server.set_debuglevel(5)

        # Make Login
        try:
            self.server.login(self.my_email, self.my_password)  # login into smpt server
        except smtplib.SMTPAuthenticationError:
            print("""
                                Your account "{}" blocked the connection, please check your mailbox and accept the terms of 
                                access from applications not secured.
                                """.format(self.my_email))


def alphabet_position(letter: str) -> int:
    """
        Returns the index of a letter passed to the function.
        This used only for openpyxl column values
    :param letter: A letter to convert to int
    :return: Return a int that is the index of the letter in the alphabet
    """

    alphabet = {'a': '1', 'b': '2', 'c': '3', 'd': '4', 'e': '5', 'f': '6', 'g': '7', 'h': '8',
                'i': '9', 'j': '10', 'k': '11', 'l': '12', 'm': '13', 'n': '14', 'o': '15', 'p': '16', 'q': '17',
                'r': '18', 's': '19', 't': '20', 'u': '21', 'v': '22', 'w': '23', 'x': '24', 'y': '25', 'z': '26'
                }
    for key, value in alphabet.items():
        if key == letter:
            return value


def email_sender(send_to: list, subject: str, message: str) -> None:
    """
           This function send emails from a sender, that is the email where the emails will be sended from,
           to a list of emails, with the same subject and message.
            This function assumes that all emails are valid.
    :param send_to: This is a list of emails, to send messages, all of this emails will recieve the same message
    :param subject: This is the subject of the message
    :param message: This is the message itself
    :return: This function returns anything
    """

    # Bind server to smpt server, running on port 587, to connect to gmail server

    for email_to_send in send_to:
        # Create the body of the message
        msg = MIMEText(message)
        msg['From'] = my_variable_globals.my_email
        msg['To'] = email_to_send
        msg['Subject'] = subject
        # Tell server to send the email
        try:
            my_variable_globals.server.sendmail(my_variable_globals.my_email, email_to_send, msg.as_string())
        except smtplib.SMTPRecipientsRefused:
            print("The recipient refused the connection.")
        except smtplib.SMTPHeloError:
            print("The server refused the Helo message.")
        except smtplib.SMTPSenderRefused:
            print("Sender address refused the messages.")
        except smtplib.SMTPDataError:
            print("The SMTP server refused to accept the message data")


def load_workbook_and_get_emails(path_to_wb: str, sheet_name: str, column_emails: str) -> list:
    """
        This function open a xlsx file and returns all the emails in a list

    :param path_to_wb: This is the path to the xlsx file. E.g: /home/mendes/workbook.xlsx
    :param sheet_name: This is the workbook sheet name, which contains the emails data
    :param column_emails: This is a letter that represent the letter of the column. E.g: A, B, C, D
    :return: This function returns a list of emails.
    """

    try:
        sheet = openpyxl.load_workbook(path_to_wb)[sheet_name]
    except KeyError:
        raise KeyError("Any sheet found with that title.")
    except InvalidFileException:
        raise InvalidFileException("Couldn't open the path expecified.")
    except SheetTitleException:
        raise SheetTitleException("Sheet title Misspelled.")

    # All rows in the xlsx file
    emails = []
    for row_count in range(2, sheet.max_row + 1):
        try:
            emails.append(sheet.cell(row=row_count, column=int(alphabet_position(column_emails.lower()))).value)
        except ValueError:
            print("Value not found in row: {}".format(row_count))
            continue

    emails = [email for email in emails if email is not None]

    return emails



if __name__ == '__main__':
    my_variable_globals = GlobalStorage()  # If needs to change email,and password, only needs to add it as argrument
    my_variable_globals.connect_to_smtp_server()
    email_sender(["a.jplmendes@esmcastilho.pt"], "Teste", "Teste")
